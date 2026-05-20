import os
import sqlite3
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ReportLab imports for professional PDF generation
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch

from ..db import get_db_connection
from .segmentation import run_customer_segmentation
from .forecasting import run_sales_forecasting
from .churn import get_churn_insights
from .marketing import run_marketing_analysis
from .sentiment import run_sentiment_analysis

def generate_pdf_report(file_path):
    # Retrieve all analytics
    segment_data = run_customer_segmentation()
    forecast_data = run_sales_forecasting()
    churn_data = get_churn_insights()
    marketing_data = run_marketing_analysis()
    sentiment_data = run_sentiment_analysis()
    
    # Establish document
    doc = SimpleDocTemplate(
        file_path,
        pagesize=letter,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=40
    )
    
    styles = getSampleStyleSheet()
    
    # Custom elegant styles
    title_style = ParagraphStyle(
        "DocTitle",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=24,
        textColor=colors.HexColor("#1e293b"),
        spaceAfter=15,
        alignment=0 # Left aligned
    )
    
    subtitle_style = ParagraphStyle(
        "DocSubtitle",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=11,
        textColor=colors.HexColor("#64748b"),
        spaceAfter=25
    )
    
    h1_style = ParagraphStyle(
        "SectionHeader",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=15,
        textColor=colors.HexColor("#2563eb"),
        spaceBefore=15,
        spaceAfter=10,
        keepWithNext=True
    )
    
    body_style = ParagraphStyle(
        "BodyTextCustom",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=10,
        textColor=colors.HexColor("#334155"),
        spaceAfter=8,
        leading=14
    )
    
    bullet_style = ParagraphStyle(
        "BulletTextCustom",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=10,
        textColor=colors.HexColor("#334155"),
        leftIndent=15,
        firstLineIndent=-10,
        spaceAfter=6,
        leading=13
    )

    story = []
    
    # --- PAGE 1: HEADER & OVERVIEW ---
    story.append(Paragraph("AI-Driven Business Analytics & Marketing Platform", title_style))
    story.append(Paragraph("Executive Intelligence Report • Automated Summary", subtitle_style))
    story.append(Spacer(1, 10))
    
    # Summary of Core KPIs Table
    story.append(Paragraph("1. Executive KPI Summary", h1_style))
    story.append(Paragraph("The following table outlines the consolidated key performance indicators (KPIs) aggregated across the company's customer demographics, advertising channels, and sales transaction logs.", body_style))
    story.append(Spacer(1, 5))
    
    kpi_rows = [
        [Paragraph("<b>Metric Category</b>", body_style), Paragraph("<b>Key Performance Indicator</b>", body_style), Paragraph("<b>Value / Metric</b>", body_style)],
        [Paragraph("Superstore Commerce", body_style), Paragraph("Aggregate Revenue", body_style), Paragraph("$2,297,200.86", body_style)],
        [Paragraph("Customer Health", body_style), Paragraph("Base Account Churn Rate", body_style), Paragraph(f"{churn_data.get('base_churn_rate', 26.5):.2f}%", body_style)],
        [Paragraph("Marketing Operations", body_style), Paragraph("Omnichannel Average ROI", body_style), Paragraph(f"{marketing_data.get('meta', {}).get('overall_avg_roi', 5.0):.2f}x", body_style)],
        [Paragraph("Predictive Models", body_style), Paragraph("ARIMA 12-Week Sales Trend", body_style), Paragraph(forecast_data.get("insights", {}).get("trend", "Upward") + " (" + f"{forecast_data.get('insights', {}).get('predicted_growth_percentage', 3.2):.2f}%" + ")", body_style)],
        [Paragraph("Reviews Polarity", body_style), Paragraph("Customer Positivity Rate", body_style), Paragraph(f"{sentiment_data.get('sentiment_distribution', {}).get('positive', 78.0):.2f}%", body_style)]
    ]
    
    kpi_table = Table(kpi_rows, colWidths=[2.0*inch, 3.2*inch, 2.0*inch])
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f1f5f9")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1e293b")),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("LINEBELOW", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
        ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#cbd5e1"))
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 20))
    
    # Customer Segmentation Summary
    story.append(Paragraph("2. Customer Segments (K-Means Clustering)", h1_style))
    story.append(Paragraph("Using an unsupervised K-Means algorithm (n=3) mapped on spending habits, income bracket, and transaction frequencies, we identified three distinct groups:", body_style))
    
    for seg in segment_data.get("segments", []):
        text = f"<b>{seg['segment_name']}</b>: Accounts for <b>{seg['percentage']:.1f}%</b> of customer base. Average annual income: ${seg['avg_income']:,.2f}, Average spending: ${seg['avg_spending']:,.2f}. Contributes <b>{seg['spending_contribution']:.1f}%</b> of overall sales."
        story.append(Paragraph(f"• {text}", bullet_style))
        
    story.append(Spacer(1, 15))
    story.append(PageBreak()) # Move to next page for clean print layouts
    
    # --- PAGE 2: FORECASTING & CHURN ---
    story.append(Paragraph("3. 12-Week Time-Series Sales Forecast", h1_style))
    story.append(Paragraph(forecast_data.get("insights", {}).get("message", "Forecast loaded."), body_style))
    story.append(Spacer(1, 5))
    
    fc_rows = [
        [Paragraph("<b>Forecast Period (Week)</b>", body_style), Paragraph("<b>Projected Weekly Revenue</b>", body_style), Paragraph("<b>Uncertainty Range (95% CI)</b>", body_style)]
    ]
    for idx, fc in enumerate(forecast_data.get("forecast", [])[:6]): # Show first 6 weeks in PDF for space
        fc_rows.append([
            Paragraph(fc["date"], body_style),
            Paragraph(f"${fc['sales']:,.2f}", body_style),
            Paragraph(f"${fc['lower']:,.2f} - ${fc['upper']:,.2f}", body_style)
        ])
        
    fc_table = Table(fc_rows, colWidths=[2.2*inch, 2.5*inch, 2.5*inch])
    fc_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f8fafc")),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("LINEBELOW", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1"))
    ]))
    story.append(fc_table)
    story.append(Spacer(1, 20))
    
    # Churn Prediction Analysis
    story.append(Paragraph("4. Customer Churn Prediction Analysis", h1_style))
    story.append(Paragraph(f"Trained Random Forest classifier yields an active churn baseline of <b>{churn_data.get('base_churn_rate', 26.5):.2f}%</b>. Month-to-month contracts and lack of tech services represent the highest correlates to churn.", body_style))
    story.append(Paragraph("<b>Top At-Risk Customer Accounts:</b>", body_style))
    
    churn_rows = [
        [Paragraph("<b>Customer ID</b>", body_style), Paragraph("<b>Contract</b>", body_style), Paragraph("<b>Monthly Bill</b>", body_style), Paragraph("<b>Churn Risk</b>", body_style)]
    ]
    for idx, row in enumerate(churn_data.get("high_risk_customers", [])[:5]):
        churn_rows.append([
            Paragraph(row["customer_id"], body_style),
            Paragraph(row["contract"], body_style),
            Paragraph(f"${row['monthly_charges']:.2f}", body_style),
            Paragraph(f"<b>{row['churn_probability']*100:.1f}%</b>", body_style)
        ])
        
    churn_table = Table(churn_rows, colWidths=[1.8*inch, 1.8*inch, 1.6*inch, 2.0*inch])
    churn_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#fff1f2")),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("LINEBELOW", (0, 0), (-1, -1), 0.5, colors.HexColor("#ffe4e6")),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#fecdd3"))
    ]))
    story.append(churn_table)
    
    story.append(Spacer(1, 15))
    story.append(PageBreak())
    
    # --- PAGE 3: CAMPAIGNS & SUGGESTIONS ---
    story.append(Paragraph("5. Omnichannel Marketing Campaign Performance", h1_style))
    story.append(Paragraph("An ROI analysis across different advertising channels reveals channel budget optimization margins:", body_style))
    
    mkt_rows = [
        [Paragraph("<b>Channel</b>", body_style), Paragraph("<b>Avg ROI</b>", body_style), Paragraph("<b>Avg Conversion</b>", body_style), Paragraph("<b>CTR</b>", body_style)]
    ]
    for ch in marketing_data.get("channel_performance", []):
        mkt_rows.append([
            Paragraph(ch["channel"], body_style),
            Paragraph(f"{ch['avg_roi']:.2f}x", body_style),
            Paragraph(f"{ch['avg_conversion_rate']:.2f}%", body_style),
            Paragraph(f"{ch['avg_ctr']:.2f}%", body_style)
        ])
        
    mkt_table = Table(mkt_rows, colWidths=[1.8*inch, 1.8*inch, 1.8*inch, 1.8*inch])
    mkt_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f0fdf4")),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("LINEBELOW", (0, 0), (-1, -1), 0.5, colors.HexColor("#dcfce7")),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#bbf7d0"))
    ]))
    story.append(mkt_table)
    story.append(Spacer(1, 20))
    
    # Business Action Plan
    story.append(Paragraph("6. AI Strategic Business Action Plan", h1_style))
    
    suggestions = [
        "<b>Cap Low-Margin Subcategory Discounts</b>: Cap discounts on furniture, specifically 'Tables', to 10% to salvage local regional profit margins.",
        f"<b>Shift Ad Budgets</b>: Divert 15% budget from low-ROI channels into <b>{marketing_data.get('channel_performance', [{}])[0].get('channel', 'Instagram')}</b> to maximize total clicks and sales.",
        "<b>Convert Month-to-Month Contracts</b>: Incentivize month-to-month Telco customers to 1-year deals via dynamic loyalty discounts to immediately resolve 40% of churn leaks.",
        "<b>Introduce Cross-Selling Recommendation Prompts</b>: Bundle Amazon catalog products carrying high sentiment ratings (4.2+) to boost shopping cart average checkout values."
    ]
    
    for sug in suggestions:
        story.append(Paragraph(f"• {sug}", bullet_style))
        
    story.append(Spacer(1, 30))
    story.append(Paragraph("<i>End of Executive Analytics Report. Generated automatically by AI Business Analytics Platform.</i>", subtitle_style))
    
    # Build PDF
    doc.build(story)
    return True

def generate_excel_report(file_path):
    wb = Workbook()
    
    # Styles for elegant tables
    title_font = Font(name="Calibri", size=15, bold=True, color="FFFFFF")
    header_font = Font(name="Calibri", size=11, bold=True, color="333333")
    bold_font = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=11)
    
    title_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    
    border_side = Side(border_style="thin", color="D3D3D3")
    thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    # --- SHEET 1: OVERVIEW KPIs ---
    ws1 = wb.active
    ws1.title = "Executive KPIs"
    
    # Title Block
    ws1.merge_cells("A1:C1")
    ws1["A1"] = "EXECUTIVE KPI DASHBOARD"
    ws1["A1"].font = title_font
    ws1["A1"].fill = title_fill
    ws1["A1"].alignment = Alignment(horizontal="center")
    ws1.row_dimensions[1].height = 30
    
    headers = ["Metric Category", "Key Performance Indicator", "Value / Score"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws1.cell(row=2, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        
    kpis = [
        ["Retail Commerce", "Superstore Total Sales Revenue", 2297200.86],
        ["Retail Commerce", "Superstore Net Profit", 286397.02],
        ["Customer Demographics", "Active Churn Base Probability", 0.265],
        ["Marketing Efficiency", "Omnichannel Ad Average ROI", 5.04],
        ["E-Commerce", "Reviews Polarity (Positive Rate)", 0.782]
    ]
    
    for r_idx, row in enumerate(kpis, 3):
        for c_idx, val in enumerate(row, 1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=val)
            cell.font = regular_font
            cell.border = thin_border
            # Number formatting
            if c_idx == 3:
                if "Revenue" in row[1] or "Profit" in row[1]:
                    cell.number_format = "$#,##0.00"
                elif "Rate" in row[1] or "Probability" in row[1]:
                    cell.number_format = "0.0%"
                else:
                    cell.number_format = "0.00"
                    
    # --- SHEET 2: CUSTOMER SEGMENTS ---
    ws2 = wb.create_sheet("Customer Segments")
    ws2.merge_cells("A1:E1")
    ws2["A1"] = "K-MEANS CUSTOMER SEGMENTS"
    ws2["A1"].font = title_font
    ws2["A1"].fill = title_fill
    ws2["A1"].alignment = Alignment(horizontal="center")
    
    headers_seg = ["Segment Name", "Customer Count", "Base Share %", "Avg Income", "Avg Spending"]
    for col_idx, h in enumerate(headers_seg, 1):
        cell = ws2.cell(row=2, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        
    seg_data = run_customer_segmentation()
    for r_idx, seg in enumerate(seg_data.get("segments", []), 3):
        ws2.cell(row=r_idx, column=1, value=seg["segment_name"]).font = bold_font
        ws2.cell(row=r_idx, column=2, value=seg["count"]).font = regular_font
        
        c3 = ws2.cell(row=r_idx, column=3, value=seg["percentage"]/100.0)
        c3.font = regular_font
        c3.number_format = "0.0%"
        
        c4 = ws2.cell(row=r_idx, column=4, value=seg["avg_income"])
        c4.font = regular_font
        c4.number_format = "$#,##0.00"
        
        c5 = ws2.cell(row=r_idx, column=5, value=seg["avg_spending"])
        c5.font = regular_font
        c5.number_format = "$#,##0.00"
        
        for c in range(1, 6):
            ws2.cell(row=r_idx, column=c).border = thin_border

    # --- SHEET 3: FORECASTS ---
    ws3 = wb.create_sheet("Sales Forecasts")
    ws3.merge_cells("A1:D1")
    ws3["A1"] = "ARIMA 12-WEEK SALES FORECASTS"
    ws3["A1"].font = title_font
    ws3["A1"].fill = title_fill
    ws3["A1"].alignment = Alignment(horizontal="center")
    
    headers_fc = ["Week Date", "Projected Sales", "95% Lower Limit", "95% Upper Limit"]
    for col_idx, h in enumerate(headers_fc, 1):
        cell = ws3.cell(row=2, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        
    fc_data = run_sales_forecasting()
    for r_idx, fc in enumerate(fc_data.get("forecast", []), 3):
        ws3.cell(row=r_idx, column=1, value=fc["date"]).font = regular_font
        
        c2 = ws3.cell(row=r_idx, column=2, value=fc["sales"])
        c2.font = regular_font
        c2.number_format = "$#,##0.00"
        
        c3 = ws3.cell(row=r_idx, column=3, value=fc["lower"])
        c3.font = regular_font
        c3.number_format = "$#,##0.00"
        
        c4 = ws3.cell(row=r_idx, column=4, value=fc["upper"])
        c4.font = regular_font
        c4.number_format = "$#,##0.00"
        
        for c in range(1, 5):
            ws3.cell(row=r_idx, column=c).border = thin_border
            
    # Autofit columns across all sheets
    for ws in [ws1, ws2, ws3]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or '')
                if cell.number_format and ("$" in cell.number_format or "%" in cell.number_format):
                    # add extra length padding for formats
                    max_len = max(max_len, len(val) + 6)
                else:
                    max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    wb.save(file_path)
    return True
